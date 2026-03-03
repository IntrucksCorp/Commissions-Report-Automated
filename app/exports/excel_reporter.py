from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import WriteOnlyCell


# Colores profesionales
HEADER_FILL = PatternFill(start_color="2E5C8A",
                          end_color="2E5C8A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Arial", size=10)

# Formato de dinero
MONEY_FORMAT = '$#,##0.00;[Red]($#,##0.00)'

# Bordes
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


def export_endorsements_to_excel(endorsements, filename, date_from="2025-12-01", date_to=None):
    """
    Exporta endorsements a Excel con 2 hojas:
    1. Endorsement Detail (Detalle completo)
    2. Agent Summary (Resumen por agente mensual)
    """
    print(f"🔹 Exportando a Excel en '{filename}' ...")

    # Convertir a lista para permitir múltiples iteraciones
    endorsements_list = list(endorsements)

    # 1. Crear Workbook
    wb = Workbook()

    # --- Hoja 1: Endorsement Detail ---
    ws_detail = wb.active
    ws_detail.title = "Endorsement Detail"
    
    # Metadata Detail
    metadata = [
        ["REPORTE DE DETALLE DE ENDORSEMENTS"],
        [f"Fecha Desde:", date_from or "N/A"],
        [f"Fecha Hasta:", date_to or "Hoy"],
        [f"Generado el:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        []
    ]
    for row in metadata:
        ws_detail.append(row)
    
    ws_detail["A1"].font = Font(bold=True, size=14)

    # Headers Detail
    headers_detail = [
        "Agent/CSR", "Endorsement Date", "Endorsement Amount",
        "Endorsement Type", "MGA", "Policy Number",
        "Policy Effective", "Policy Expiration", "Insured",
        "Agency Commission", "Agent Commission"
    ]
    ws_detail.append(headers_detail)
    header_row_detail = len(metadata) + 1

    for col_idx, header in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=header_row_detail, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws_detail.row_dimensions[header_row_detail].height = 35

    # Sort data by agent for grouping
    endorsements_list.sort(key=lambda x: (x.get("agent") or "").lower())

    # Contenido Detail
    detail_count = 0
    detail_total_agency_comm = 0
    detail_total_agent_comm = 0
    
    current_agent = None
    agent_agency_sum = 0
    agent_agent_sum = 0
    
    SUBTOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    SUBTOTAL_FONT = Font(bold=True, name="Arial", size=10)

    def append_subtotal_row(ws, agent_name, agency_sum, agent_sum, is_grand_total=False):
        label = "GRAND TOTAL" if is_grand_total else f"Total {agent_name or 'N/A'}"
        ws.append([
            label,
            None, None, None, None, None, None, None, None,
            agency_sum,
            agent_sum
        ])
        row_num = ws.max_row
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = SUBTOTAL_FONT
            cell.fill = SUBTOTAL_FILL if not is_grand_total else PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = THIN_BORDER
            if col_idx in [10, 11]:
                cell.number_format = MONEY_FORMAT
                val = agency_sum if col_idx == 10 else agent_sum
                if val < 0:
                    cell.font = Font(bold=True, name="Arial", size=10, color="FF0000")
        ws.row_dimensions[row_num].height = 20

    for e in endorsements_list:
        agent = e.get("agent")
        
        # Detect agent change for subtotal
        if current_agent is not None and agent != current_agent:
            append_subtotal_row(ws_detail, current_agent, agent_agency_sum, agent_agent_sum)
            agent_agency_sum = 0
            agent_agent_sum = 0
        
        current_agent = agent
        
        endorsement_type_raw = e.get("endorsement_type") or ""
        is_cancel = "cancel" in endorsement_type_raw.lower()

        amount = safe_money(e.get("endorsement_amount"))
        if is_cancel and amount > 0:
            amount = -amount

        agency_comm = safe_money(e.get("agency_commission"))
        agent_comm = safe_money(e.get("agent_commission"))

        if is_cancel:
            agency_comm = -abs(agency_comm) if agency_comm != 0 else 0
            agent_comm = -abs(agent_comm) if agent_comm != 0 else 0

        # Accumulate totals
        agent_agency_sum += agency_comm
        agent_agent_sum += agent_comm
        
        detail_total_agency_comm += agency_comm
        detail_total_agent_comm += agent_comm
        detail_count += 1

        ws_detail.append([
            e.get("agent"),
            _format_date(e.get("endorsement_effective")),
            amount,
            endorsement_type_raw,
            e.get("mga"),
            e.get("policy_number"),
            _format_date(e.get("policy_effective_date")),
            _format_date(e.get("policy_expiration_date")),
            e.get("insured"),
            agency_comm,
            agent_comm
        ])

        curr_row = ws_detail.max_row

        for col_idx in range(1, 12):
            cell = ws_detail.cell(row=curr_row, column=col_idx)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx in [3, 10, 11]:
                cell.number_format = MONEY_FORMAT
                if (col_idx == 3 and amount < 0) or (col_idx == 10 and agency_comm < 0) or (col_idx == 11 and agent_comm < 0):
                    cell.font = Font(name="Arial", size=10, color="FF0000")

        ws_detail.row_dimensions[curr_row].height = 20

    # Final subtotals
    if current_agent is not None or detail_count > 0:
        append_subtotal_row(ws_detail, current_agent, agent_agency_sum, agent_agent_sum)
        
        # Grand Total Row
        append_subtotal_row(ws_detail, None, detail_total_agency_comm, detail_total_agent_comm, is_grand_total=True)

    # Column Widths Detail
    detail_widths = [30, 16, 18, 26, 32, 20, 15, 15, 30, 18, 18]
    for i, w in enumerate(detail_widths, 1):
        ws_detail.column_dimensions[chr(64 + i)].width = w

    ws_detail.freeze_panes = "A" + str(header_row_detail + 1)
    ws_detail.auto_filter.ref = f"A{header_row_detail}:K{ws_detail.max_row}"

    print(f"💾 Hoja 'Endorsement Detail' creada: {detail_count} filas")

    # --- Hoja 2: Agent Summary (MENSUAL) ---
    ws_summary = wb.create_sheet("Agent Summary")
    
    # Procesar datos para resumen por agente Y MES
    agent_summary_dict = {}
    
    for e in endorsements_list:
        agent = e.get("agent")
        if not agent:
            continue
        
        # Extraer fecha del endorsement para agrupar por mes
        eff_date_raw = e.get("endorsement_effective")
        try:
            # "2025-12-13T00:00:00" -> "December", "2025"
            date_str = str(eff_date_raw).split("T")[0]
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            m_name = date_obj.strftime("%B")
            y_val = date_obj.strftime("%Y")
            sort_key_date = date_obj.strftime("%Y%m") # Para ordenamiento cronológico
        except:
            m_name = "N/A"
            y_val = "N/A"
            sort_key_date = "000000"

        agent_comm = safe_money(e.get("agent_commission"))
        
        # Detectar si es cancelación para el signo
        endorsement_type_raw = e.get("endorsement_type") or ""
        if "cancel" in endorsement_type_raw.lower() and agent_comm > 0:
             agent_comm = -agent_comm
        elif "cancel" in endorsement_type_raw.lower():
             agent_comm = -abs(agent_comm) if agent_comm != 0 else 0

        # La clave de agrupación es (agente, año, mes)
        group_key = (agent, y_val, m_name)
        if group_key not in agent_summary_dict:
            agent_summary_dict[group_key] = {
                "count": 0,
                "total_agent_comm": 0.0,
                "sort_date": sort_key_date
            }
        
        agent_summary_dict[group_key]["count"] += 1
        agent_summary_dict[group_key]["total_agent_comm"] += agent_comm

    # Filtrar y preparar para ordenar
    filtered_groups = []
    for key, data in agent_summary_dict.items():
        if abs(data["total_agent_comm"]) > 0.001:
            filtered_groups.append({
                "agent": key[0],
                "year": key[1],
                "month": key[2],
                "count": data["count"],
                "total": data["total_agent_comm"],
                "sort_date": data["sort_date"]
            })

    # Ordenar: Fecha DESC (sort_date), luego Total DESC, luego Agent ASC
    filtered_groups.sort(key=lambda x: (x["sort_date"], -x["total"], x["agent"]), reverse=False)
    # Queremos los meses más recientes arriba, así que invertimos el sort_date en el key o invertimos el resultado
    filtered_groups.sort(key=lambda x: (x["sort_date"], x["total"]), reverse=True)
    # Mejor así: Fecha DESC, luego dentro de ese mes por Total DESC
    filtered_groups.sort(key=lambda x: x["sort_date"], reverse=True)
    
    # Re-ordenar final: Meses recientes primero, y dentro de cada mes, mayor comisión primero
    filtered_groups.sort(key=lambda x: (x["sort_date"], x["total"]), reverse=True)

    summary_headers = ["Agent/CSR", "Month", "Year", "Total Endorsements", "Total Agent Commission"]
    ws_summary.append(summary_headers)
    
    # Formato Headers Summary
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws_summary.row_dimensions[1].height = 35

    # Datos Summary
    row_idx = 2
    summary_total_count = 0
    summary_total_comm = 0
    
    for gdata in filtered_groups:
        ws_summary.append([
            gdata["agent"],
            gdata["month"],
            gdata["year"],
            gdata["count"],
            gdata["total"]
        ])
        
        # Estilos celdas
        for col_idx in range(1, 6):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            
            if col_idx == 1: cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [2, 3, 4]: cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = MONEY_FORMAT
                if gdata["total"] < 0:
                     cell.font = Font(name="Arial", size=10, color="FF0000")
        
        summary_total_count += gdata["count"]
        summary_total_comm += gdata["total"]
        ws_summary.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Fila de Totales Summary
    ws_summary.append(["TOTAL", "", "", summary_total_count, summary_total_comm])
    last_row_summary = ws_summary.max_row
    
    TOTAL_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    DOUBLE_BORDER_TOP = Border(
        top=Side(style='double', color='000000'),
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for col_idx in range(1, 6):
        cell = ws_summary.cell(row=last_row_summary, column=col_idx)
        cell.font = Font(bold=True, name="Arial", size=10)
        cell.fill = TOTAL_FILL
        cell.border = DOUBLE_BORDER_TOP
        if col_idx == 4: cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 5:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = MONEY_FORMAT
            if summary_total_comm < 0:
                 cell.font = Font(bold=True, name="Arial", size=10, color="FF0000")

    # Column Widths Summary
    summary_widths = [30, 15, 10, 18, 22]
    for i, w in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[chr(64 + i)].width = w

    ws_summary.freeze_panes = "B2"
    ws_summary.auto_filter.ref = f"A1:E{last_row_summary-1}"

    print(f"💾 Hoja 'Agent Summary' creada: {len(filtered_groups)} filas (Mensual)")

    # 4. Validar Totales
    assert abs(summary_total_comm - detail_total_agent_comm) < 0.01, f"Error: Totales no coinciden. Summary: {summary_total_comm}, Detail: {detail_total_agent_comm}"
    print("✅ Totales validados correctamente")

    # 5. Guardar
    wb.save(filename)
    print(f"✅ Excel generado: {filename} (Total: {detail_count} filas)")


# -----------------------
# Helpers
# -----------------------

def _format_date(value):
    """Formatea fechas ISO a formato MM/DD/YYYY."""
    if not value:
        return None
    try:
        # Si viene como "2025-12-13T00:00:00" o "2025-12-13"
        date_str = str(value).split("T")[0]
        # Separar año-mes-día
        parts = date_str.split("-")
        if len(parts) == 3:
            year, month, day = parts
            return f"{month}/{day}/{year}"
        return value
    except:
        return value


def safe_money(value):
    """Convierte valores a float de forma segura."""
    try:
        if value is None:
            return 0.0
        return float(value)
    except:
        return 0.0
